import openpyxl
import sqlite3

connection = sqlite3.connect('airbnb.db')
cursor = connection.cursor()

neighborHoodWorkbook = openpyxl.load_workbook('neighbourhoods.xlsx')
for row in list(neighborHoodWorkbook.active.values)[1:]:
    cursor.execute('INSERT into neighborhoods(name) values(?)', (row[1],))

calendarWorkbook = openpyxl.load_workbook('calendar.xlsx')
for row in list(calendarWorkbook.active.values)[1:]:
    cursor.execute('INSERT into calendar(listing_id, date, available, price,\
                   adjusted_price, minimum_nights, maximum_nights)\
                   values(?, ?, ?, ?, ?, ?, ?)',\
                   (row[0],row[1], row[2], row[3], row[4], row[5], row[6]))

reviewsWorkbook = openpyxl.load_workbook('reviews.xlsx')
for row in list(reviewsWorkbook.active.values)[1:]:
    cursor.execute('INSERT into reviews(id, listing_id, date, reviewer_name,\
                   reviewer_id, comments) values(?, ?, ?, ?, ?, ?)',\
                   (row[1],row[0], row[2], row[4], row[3], row[5]))

listingsWorkbook = openpyxl.load_workbook('listings.xlsx')
for row in list(listingsWorkbook.active.values)[1:]:
    cursor.execute('INSERT into listings(id, listing_url, scrape_id,\
                   last_scraped, source, name, description,\
                   neighborhood_overview, picture_url, host_id, host_url,\
                   host_name, host_since, host_location, host_about,\
                   host_response_time, host_response_rate, host_acceptance_rate,\
                   host_is_superhost, host_thumbnail_url, host_picture_url,\
                   host_neighbourhood, host_listings_count,\
                   host_total_listings_count, host_verifications,\
                   host_has_profile_pic, host_identity_verified, neighbourhood,\
                   neighbourhood_cleansed, neighbourhood_group_cleansed,\
                   latitude, longitude, property_type, room_type, accommodates,\
                   bathrooms, bathrooms_text, bedrooms, beds, amenities, price,\
                   minimum_nights, maximum_nights, minimum_minimum_nights,\
                   maximum_minimum_nights, minimum_maximum_nights,\
                   maximum_maximum_nights, minimum_nights_avg_ntm,\
                   maximum_nights_avg_ntm, calendar_updated, has_availability,\
                   availability_30, availability_60, availability_90,\
                   availability_365, calendar_last_scraped, number_of_reviews,\
                   number_of_reviews_ltm, number_of_reviews_l30d, first_review,\
                   last_review, review_scores_rating, review_scores_accuracy,\
                   review_scores_cleanliness, review_scores_checkin,\
                   review_scores_communication, review_scores_location,\
                   review_scores_value, license, instant_bookable,\
                   calculated_host_listings_count,\
                   calculated_host_listings_count_entire_homes,\
                   calculated_host_listings_count_private_rooms,\
                   calculated_host_listings_count_shared_rooms, reviews_per_month)\
                   values(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,\
                   ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,\
                   ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,\
                   ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                   (row[0],row[1], row[2], row[3], row[4], row[5], row[6], row[7],\
                    row[8], row[9], row[10],row[11], row[12], row[13], row[14],\
                    row[15], row[16], row[17], row[18], row[19], row[20],row[21],\
                    row[22], row[23], row[24], row[25], row[26], row[27], row[28],\
                    row[29], row[30], row[31], row[32], row[33], row[34], row[35],\
                    row[36], row[37], row[38], row[39], row[40], row[41], row[42],\
                    row[43], row[44], row[45], row[46], row[47], row[48], row[49],\
                    row[50], row[51], row[52], row[53], row[54], row[55], row[56],\
                    row[57], row[58], row[59], row[60], row[61], row[62], row[63],\
                    row[64], row[65], row[66], row[67], row[68], row[69], row[70],\
                    row[71], row[72], row[73], row[74]))

connection.commit()
connection.close()
